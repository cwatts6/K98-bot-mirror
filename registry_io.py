# registry_io.py
"""
Registry I/O, validation, CSV and XLSX helpers.

This module centralizes:
 - exports (generic CSV + richer audit CSV and XLSX)
 - parsing CSV/XLSX bytes into rows (robust to Excel BOM and formula cells)
 - normalization of account types and GovernorIDs (GovernorID must be numeric)
 - preparing an import plan (validations, in-file duplicate checks, conflict checks)
 - applying an import plan (atomic save via governor_registry.save_registry)
 - building error CSV bytes for reporting invalid rows

XLSX support uses openpyxl. If openpyxl is not installed, XLSX helpers will raise ImportError.
"""

from __future__ import annotations

import copy
import csv
from decimal import Decimal, InvalidOperation
import io
import logging
import re

import governor_registry
from utils import normalize_governor_id

logger = logging.getLogger(__name__)

__all__ = [
    "CSV_HEADERS",
    "_strip_excel_formula",
    "apply_import_plan",
    "build_error_csv_bytes",
    "export_registration_audit_files",
    "export_registration_audit_xlsx_bytes",
    "export_registry_to_csv_bytes",
    "export_registry_to_xlsx_bytes",
    "normalize_account_type",
    "normalize_id",
    "parse_csv_bytes",
    "parse_xlsx_bytes",
    "prepare_import_plan",
    "rows_to_xlsx_bytes",
]

CSV_HEADERS = ["DiscordUserID", "AccountType", "GovernorID", "GovernorName"]

# Regexes for normalization
_SCI_RE = re.compile(r"^\s*[+-]?\d+(?:\.\d+)?[eE][+-]?\d+\s*$")
_DOT_ZERO_RE = re.compile(r"\.0+$")
_DIGITS_RE = re.compile(r"^\d+$")


# ---------- CSV helpers ----------
def export_registry_to_csv_bytes(registry: dict) -> io.BytesIO:
    """
    Generic CSV export of the registry (no roles, no Excel-safe dual columns).
    """
    buf = io.StringIO(newline="")
    writer = csv.DictWriter(buf, fieldnames=CSV_HEADERS, lineterminator="\n")
    writer.writeheader()
    for discord_id, user_block in (registry or {}).items():
        accounts = (user_block or {}).get("accounts", {}) or {}
        for acct_label, acct in accounts.items():
            writer.writerow(
                {
                    "DiscordUserID": str(discord_id),
                    "AccountType": str(acct_label),
                    "GovernorID": str(acct.get("GovernorID", "")),
                    "GovernorName": str(acct.get("GovernorName", "")),
                }
            )
    out = io.BytesIO()
    out.write(buf.getvalue().encode("utf-8"))
    out.seek(0)
    return out


def parse_csv_bytes(content: bytes, encoding: str = "utf-8") -> list[dict[str, str]]:
    """
    Parse CSV content (bytes) into list of dict rows using csv.DictReader.
    Decoding uses 'encoding' with errors='replace' to be tolerant.
    """
    text = content.decode(encoding, errors="replace")
    f = io.StringIO(text)
    reader = csv.DictReader(f)
    rows: list[dict[str, str]] = []
    for r in reader:
        normalized_row = {
            str(k).strip(): (str(v).strip() if v is not None else "") for k, v in r.items()
        }
        rows.append(normalized_row)
    return rows


# ---------- XLSX helpers (optional dependency on openpyxl) ----------
def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except Exception as e:
        raise ImportError(
            "XLSX helper requires openpyxl. Install with `pip install openpyxl`."
        ) from e


def export_registry_to_xlsx_bytes(
    registry: dict, *, sheet_name: str = "registrations"
) -> io.BytesIO:
    """
    Export the registry to a single-sheet XLSX workbook and return BytesIO.
    GovernorID and DiscordUserID columns are written as text (number_format='@') to avoid Excel coercion.
    """
    _ensure_openpyxl()
    from openpyxl import Workbook
    from openpyxl.styles import Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = [
        "discord_id",
        "discord_id_excel",
        "discord_user",
        "account_type",
        "governor_id",
        "governor_id_excel",
        "governor_name",
        "roles",
        "top_role",
    ]
    ws.append(headers)

    for discord_id, user_block in (registry or {}).items():
        accounts = (user_block or {}).get("accounts", {}) or {}
        for acc_label, acct in accounts.items():
            discord_id_s = str(discord_id).strip()
            gov_id_raw = str(acct.get("GovernorID", "")).strip()
            row = [
                discord_id_s,
                f'="{discord_id_s}"' if discord_id_s else "",
                user_block.get("discord_name", discord_id_s),
                str(acc_label),
                gov_id_raw,
                f'="{gov_id_raw}"' if gov_id_raw else "",
                acct.get("GovernorName", ""),
                "",  # roles not populated here
                "",
            ]
            ws.append(row)

    # Force text format for ID columns by setting number_format='@' on those cells
    # columns (1-based): discord_id (A), discord_id_excel (B), governor_id (E), governor_id_excel (F)
    text_columns = [1, 2, 5, 6]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in text_columns:
            cell = row[col_idx - 1]
            # ensure value is string and force text format
            if cell.value is None:
                continue
            cell.value = str(cell.value)
            try:
                cell.number_format = "@"
                cell.alignment = Alignment(horizontal="left")
            except Exception:
                # ignore formatting errors on older openpyxl versions
                pass

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def parse_xlsx_bytes(content: bytes, *, sheet_index: int = 0) -> list[dict[str, str]]:
    """
    Parse the first (or given) sheet of an XLSX workbook into a list of row dicts,
    using the first row as header names. All cell values are converted to strings.
    """
    _ensure_openpyxl()
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[sheet_index]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out_rows: list[dict[str, str]] = []
    for r in rows[1:]:
        d = {}
        for idx, h in enumerate(headers):
            v = r[idx] if idx < len(r) else None
            d[h] = "" if v is None else str(v).strip()
        out_rows.append(d)
    return out_rows


# Add the following function near the other XLSX helpers in registry_io.py
def rows_to_xlsx_bytes(
    rows: list[dict[str, object]], headers: list[str], *, sheet_name: str = "sheet"
) -> io.BytesIO:
    """
    Convert a list of dict rows (each mapping header->value) into an XLSX BytesIO.
    headers defines the column order. This preserves values such as 'roles' and 'top_role'.
    Sets text number_format for known ID columns to avoid Excel coercion.
    """
    _ensure_openpyxl()
    from openpyxl import Workbook
    from openpyxl.styles import Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(headers)
    for r in rows:
        row_vals = [r.get(h, "") for h in headers]
        ws.append(row_vals)

    # Determine columns that should be forced to text
    lower_headers = [h.lower() for h in headers]
    text_cols = set()
    for idx, h in enumerate(lower_headers, start=1):
        if h in ("discord_id", "discord_id_excel", "governor_id", "governor_id_excel"):
            text_cols.add(idx)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in text_cols:
            cell = row[col_idx - 1]
            if cell.value is None:
                continue
            cell.value = str(cell.value)
            try:
                cell.number_format = "@"
                cell.alignment = Alignment(horizontal="left")
            except Exception:
                pass

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ---------- normalization helpers ----------
def _strip_excel_formula(val: str | None) -> tuple[str, bool]:
    """
    Accept Excel-safe wrappers and formula-like values:
      - '="12345"' -> '12345'
      - '=12345' or '+12345' -> strip leading char
    """
    if val is None:
        return "", False
    s = str(val).strip()
    if not s:
        return "", False
    if len(s) >= 3 and s.startswith('="') and s.endswith('"'):
        return s[2:-1], True
    if s.startswith("=") or s.startswith("+"):
        return s[1:].strip(), True
    return s, False


def normalize_account_type(s: str | None) -> tuple[str | None, bool, str | None]:
    if s is None:
        return None, False, "Missing account type"
    raw = str(s).strip()
    if not raw:
        return None, False, "Empty account type"
    low = raw.lower().strip()
    if low in ("main", "m"):
        return "Main", (raw != "Main"), None
    m = re.match(r"^alt[\s\-_]*0*([1-9]\d*)$", low, re.IGNORECASE)
    if m:
        n = int(m.group(1))
        return f"Alt {n}", True, None
    m = re.match(r"^farm[\s\-_]*0*([1-9]\d*)$", low, re.IGNORECASE)
    if m:
        n = int(m.group(1))
        return f"Farm {n}", True, None
    if low.isdigit():
        return f"Alt {int(low)}", True, None
    if low in ("alternate", "alternate account"):
        return "Alt 1", True, None
    return None, False, f"Unrecognized account_type '{s}'. Use Main, Alt N, or Farm N."


def normalize_id(value: str | None) -> tuple[str | None, list[str]]:
    """
    Normalize a GovernorID candidate value (GovernorID must be numeric).

    Returns normalized_id or None, and a list of warnings (if any).

    Process:
    - Strip Excel wrappers (='...', ="...", leading '=' or '+') and wrapping quotes
    - Remove commas and spaces
    - Convert scientific notation and trailing .0 to integer where unambiguous
    - Ensure resulting string is digits-only; reject otherwise
    - Strip leading zeros (but return "0" if that's all that remains)
    - Attempt to canonicalize with utils.normalize_governor_id and fallback to digits
    """
    warnings: list[str] = []
    if not value:
        warnings.append("GovernorID missing")
        return None, warnings

    s, changed = _strip_excel_formula(value)
    if changed:
        warnings.append("formula")

    # Remove wrapping quotes if present
    if len(s) >= 2 and s[0] == '"' and s[-1] == '"':
        s = s[1:-1]
        warnings.append("quotes")

    # Remove commas/spaces
    if "," in s:
        s = s.replace(",", "")
        warnings.append("commas")
    if " " in s:
        s = s.replace(" ", "")
        warnings.append("spaces")

    # Scientific notation -> integer string if possible
    if _SCI_RE.match(s):
        try:
            s = f"{int(Decimal(s))}"
            warnings.append("sci")
        except (InvalidOperation, ValueError):
            pass

    # Strip trailing .0
    if _DOT_ZERO_RE.search(s):
        s = _DOT_ZERO_RE.sub("", s)
        warnings.append("dotzero")

    # Now require digits-only
    if not _DIGITS_RE.fullmatch(s):
        warnings.append(f"invalid_format:{value}")
        return None, warnings

    # Strip leading zeros (retain "0" if empty)
    s_norm = s.lstrip("0") or "0"

    # Try canonical normalizer first (shared util) but fallback to digits if it fails
    try:
        norm = normalize_governor_id(s_norm)
    except Exception:
        norm = None

    if norm:
        return norm, warnings

    return s_norm, warnings


# ---------- import/validation pipeline ----------
def prepare_import_plan(
    rows: list[dict[str, str]],
    existing_registry: dict,
    *,
    valid_ids: dict[str, dict] | None = None,
) -> tuple[list[dict], list[str], list[str], list[dict[str, str]]]:
    """
    Validate CSV/XLSX rows and prepare a list of changes to apply.
    Returns (changes, errors, warnings, error_rows).
    """
    changes: list[dict] = []
    errors: list[str] = []
    warnings: list[str] = []
    error_rows: list[dict[str, str]] = []

    existing_gid_to_owner: dict[str, tuple[str, str]] = {}
    for uid0, data0 in (existing_registry or {}).items():
        for at0, det0 in (data0.get("accounts") or {}).items():
            g0 = str(det0.get("GovernorID", "")).strip()
            if g0:
                existing_gid_to_owner[g0] = (str(uid0), str(at0))

    seen_combo: dict[tuple[str, str], str] = {}
    seen_gid_infile: dict[str, str] = {}

    for idx, row in enumerate(rows, start=2):

        def get_field(*keys):
            for k in keys:
                if k in row and row[k] is not None:
                    return row[k]
            return ""

        discord_user = get_field(
            "discord_id", "discord_id_excel", "DiscordUserID", "DiscordUser", "UserID"
        )
        acct_raw = get_field("account_type", "AccountType", "Account")
        gid_raw = get_field("governor_id", "governor_id_excel", "GovernorID", "GovID", "Governor")
        gname_raw = get_field("governor_name", "GovernorName", "Name")

        if not discord_user:
            msg = f"Row {idx}: Missing discord_id."
            errors.append(msg)
            error_rows.append(
                {
                    "row": idx,
                    "discord_id": "",
                    "account_type": acct_raw,
                    "governor_id": gid_raw,
                    "error": "Missing discord_id",
                }
            )
            continue

        acct_norm, acct_canon, acct_err = normalize_account_type(acct_raw)
        if acct_err:
            msg = f"Row {idx}: {acct_err}"
            errors.append(msg)
            error_rows.append(
                {
                    "row": idx,
                    "discord_id": discord_user,
                    "account_type": acct_raw,
                    "governor_id": gid_raw,
                    "error": acct_err,
                }
            )
            continue

        gid_norm, gid_warnings = normalize_id(gid_raw)
        for w in gid_warnings:
            warnings.append(f"Row {idx}: {w}")
        if not gid_norm:
            msg = f"Row {idx}: Invalid GovernorID '{gid_raw}'."
            errors.append(msg)
            error_rows.append(
                {
                    "row": idx,
                    "discord_id": discord_user,
                    "account_type": acct_norm or acct_raw,
                    "governor_id": gid_raw,
                    "error": "Invalid GovernorID after normalization",
                }
            )
            continue

        if gid_norm in seen_gid_infile and seen_gid_infile[gid_norm] != str(discord_user):
            msg = f"Row {idx}: governor_id `{gid_norm}` assigned to multiple discord_ids in CSV ({seen_gid_infile[gid_norm]} and {discord_user})."
            errors.append(msg)
            error_rows.append(
                {
                    "row": idx,
                    "discord_id": discord_user,
                    "account_type": acct_norm,
                    "governor_id": gid_norm,
                    "error": "GovernorID appears multiple times for different users in this CSV",
                }
            )
            continue
        seen_gid_infile[gid_norm] = str(discord_user)

        key = (str(discord_user), acct_norm)
        prev_gid_same_combo = seen_combo.get(key)
        if prev_gid_same_combo and prev_gid_same_combo != gid_norm:
            msg = f"Row {idx}: duplicate account row for {discord_user} {acct_norm} with conflicting governor_id ({prev_gid_same_combo} vs {gid_norm})."
            errors.append(msg)
            error_rows.append(
                {
                    "row": idx,
                    "discord_id": discord_user,
                    "account_type": acct_norm,
                    "governor_id": gid_norm,
                    "error": "Conflicting duplicate of (discord_id, account_type) in CSV",
                }
            )
            continue
        seen_combo[key] = gid_norm

        owner = existing_gid_to_owner.get(gid_norm)
        if owner and owner[0] != str(discord_user):
            existing_owner_uid, existing_acc = owner
            msg = f"Row {idx}: governor_id `{gid_norm}` already registered to {existing_owner_uid} ({existing_acc})."
            errors.append(msg)
            error_rows.append(
                {
                    "row": idx,
                    "discord_id": discord_user,
                    "account_type": acct_norm,
                    "governor_id": gid_norm,
                    "error": "GovernorID already registered to another user",
                }
            )
            continue

        if valid_ids is not None:
            if str(gid_norm) not in valid_ids:
                msg = f"Row {idx}: governor_id `{gid_norm}` not found in source data."
                errors.append(msg)
                error_rows.append(
                    {
                        "row": idx,
                        "discord_id": discord_user,
                        "account_type": acct_norm,
                        "governor_id": gid_norm,
                        "error": "GovernorID not found in source data",
                    }
                )
                continue

        changes.append(
            {
                "discord_id": str(discord_user),
                "account_type": acct_norm,
                "governor_id": gid_norm,
                "governor_name": str(gname_raw or "") or "",
                "source_row": idx,
            }
        )

    return changes, errors, warnings, error_rows


def build_error_csv_bytes(
    error_rows: list[dict[str, str]], *, headers: list[str] | None = None
) -> io.BytesIO:
    headers = headers or ["row", "discord_id", "account_type", "governor_id", "error"]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers, lineterminator="\n")
    writer.writeheader()
    for er in error_rows:
        writer.writerow({h: er.get(h, "") for h in headers})
    out = io.BytesIO(buf.getvalue().encode("utf-8-sig"))
    out.seek(0)
    return out


def build_error_xlsx_bytes(
    error_rows: list[dict[str, str]],
    *,
    headers: list[str] | None = None,
    sheet_name: str = "import_errors",
) -> io.BytesIO:
    """
    Build an XLSX workbook containing error_rows for easy viewing in Excel.
    error_rows is a list of dicts with keys matching headers (default: row, discord_id, account_type, governor_id, error).
    Returns BytesIO of the workbook.
    """
    _ensure_openpyxl()
    from openpyxl import Workbook
    from openpyxl.styles import Alignment

    headers = headers or ["row", "discord_id", "account_type", "governor_id", "error"]
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write header row
    ws.append(headers)

    for er in error_rows:
        row_vals = [er.get(h, "") for h in headers]
        ws.append(row_vals)

    # Force text format for id columns where applicable
    lower_headers = [h.lower() for h in headers]
    text_cols = set()
    for idx, h in enumerate(lower_headers, start=1):
        if h in ("discord_id", "governor_id"):
            text_cols.add(idx)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in text_cols:
            cell = row[col_idx - 1]
            if cell.value is None:
                continue
            cell.value = str(cell.value)
            try:
                cell.number_format = "@"
                cell.alignment = Alignment(horizontal="left")
            except Exception:
                pass

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def apply_import_plan(
    changes: list[dict], existing_registry: dict, *, dry_run: bool = True
) -> tuple[dict, list[str]]:
    reg_copy = copy.deepcopy(existing_registry or {})
    summary: list[str] = []
    for ch in changes:
        did = str(ch["discord_id"])
        acct = ch["account_type"]
        gid = ch["governor_id"]
        gname = ch.get("governor_name", "")
        if did not in reg_copy:
            reg_copy[did] = {"discord_name": did, "accounts": {}}
        accounts = reg_copy[did].get("accounts") or {}
        accounts[acct] = {"GovernorID": gid, "GovernorName": gname}
        reg_copy[did]["accounts"] = accounts
        summary.append(f"{did}: {acct} -> {gid} ({gname})")
    if not dry_run:
        governor_registry.save_registry(reg_copy)
        logger.info("Saved registry with %d changes", len(changes))
    return reg_copy, summary


# ---------- registration_audit exports (CSV and XLSX variants) ----------
def export_registration_audit_files(
    registry: dict,
    members_info: dict[str, dict[str, str]],
    sql_rows: list[dict[str, object]],
) -> dict[str, io.BytesIO]:
    """
    Return the three CSV files used by registration_audit as a dict of BytesIOs.
    Keys: registered_accounts.csv, unregistered_current_governors.csv, members_without_registration.csv
    (same output format as the earlier Commands.py implementation)
    """

    # reuse the CSV-building logic from earlier implementation
    def excel_safe_formula(value: str) -> str:
        v = (value or "").strip()
        return f'="{v}"' if v else ""

    def to_csv_bytes(headers: list[str], rows: list[dict[str, object]]) -> io.BytesIO:
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=headers, lineterminator="\n")
        writer.writeheader()
        for r in rows:
            writer.writerow({h: r.get(h, "") for h in headers})
        return io.BytesIO(buf.getvalue().encode("utf-8-sig"))

    def _norm_gid(val) -> str:
        if val is None:
            return ""
        if isinstance(val, int):
            s = str(val)
        elif isinstance(val, float):
            try:
                s = str(int(Decimal(str(val)).to_integral_value()))
            except Exception:
                s = str(int(val))
        elif isinstance(val, Decimal):
            s = str(int(val.to_integral_value()))
        else:
            s = str(val).strip()
            if s.startswith('="') and s.endswith('"') and len(s) >= 3:
                s = s[2:-1]
            s = s.replace(",", "")
            if re.fullmatch(r"[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?", s):
                try:
                    s = str(int(Decimal(s).to_integral_value()))
                except Exception:
                    pass
        digits = re.findall(r"\d+", s)
        if digits:
            s = "".join(digits)
        return s.lstrip("0") or ("0" if s else "")

    def _extract_gov_id(details: dict) -> str:
        if not isinstance(details, dict):
            return ""
        for k in (
            "GovernorID",
            "Governor Id",
            "GovernorId",
            "gov_id",
            "govid",
            "GovID",
            "Gov Id",
        ):
            if details.get(k):
                return str(details[k])
        for k, v in details.items():
            nk = re.sub(r"[^a-z0-9]", "", str(k).lower())
            if (
                ("governor" in nk and "id" in nk) or nk in ("govid", "govidnumber", "governorid")
            ) and v:
                return str(v)
        return ""

    # Build registered rows and set of normalized registered IDs
    registered_ids: set[str] = set()
    registered_rows: list[dict[str, object]] = []
    for uid, data in (registry or {}).items():
        uid_str = str(uid).strip()
        accounts = data.get("accounts", {}) or {}
        for acc_type, details in accounts.items():
            gov_id_raw = _extract_gov_id(details)
            gov_id_norm = _norm_gid(gov_id_raw)
            gov_name = str(details.get("GovernorName") or "Unknown").strip()
            if gov_id_norm:
                registered_ids.add(gov_id_norm)
            member_info = members_info.get(uid_str, {})
            registered_rows.append(
                {
                    "discord_id": uid_str,
                    "discord_id_excel": excel_safe_formula(uid_str),
                    "discord_user": data.get("discord_name", uid_str),
                    "account_type": str(acc_type),
                    "governor_id": str(gov_id_raw or ""),
                    "governor_id_excel": excel_safe_formula(str(gov_id_raw or "")),
                    "governor_name": gov_name,
                    "roles": member_info.get("roles", ""),
                    "top_role": member_info.get("top_role", ""),
                }
            )

    # Resolve current IDs from SQL
    current_ids: set[str] = set()
    row_by_id: dict[str, dict] = {}
    for r in sql_rows or []:
        gid_val = r.get("GovernorID")
        if gid_val is None:
            continue
        try:
            gid_sql = str(int(gid_val))
        except Exception:
            gid_sql = str(gid_val)
        current_ids.add(gid_sql)
        if gid_sql not in row_by_id:
            row_by_id[gid_sql] = r

    unregistered_ids = sorted(current_ids - registered_ids)

    unreg_rows: list[dict[str, object]] = []
    for gid_sql in unregistered_ids:
        rr = row_by_id.get(gid_sql, {})
        unreg_rows.append(
            {
                "governor_id": gid_sql,
                "governor_id_excel": excel_safe_formula(gid_sql),
                "governor_name": str(rr.get("GovernorName", "Unknown")),
                "alliance": str(rr.get("Alliance") or ""),
                "power_rank": rr.get("PowerRank", ""),
                "power": rr.get("Power", ""),
                "kill_points": rr.get("KillPoints", ""),
                "deads": rr.get("Deads", ""),
                "city_hall": rr.get("City Hall", ""),
                "location": rr.get("LOCATION", ""),
                "scan_date": rr.get("ScanDate", ""),
            }
        )

    registered_user_ids = set(str(k).strip() for k in registry.keys())
    mw_rows: list[dict[str, object]] = []
    for uid, minfo in (members_info or {}).items():
        if str(uid) not in registered_user_ids:
            mw_rows.append(
                {
                    "discord_id": str(uid),
                    "discord_id_excel": excel_safe_formula(str(uid)),
                    "discord_user": minfo.get("discord_user", ""),
                    "roles": minfo.get("roles", ""),
                    "top_role": minfo.get("top_role", ""),
                }
            )

    # Sorting
    registered_rows.sort(
        key=lambda r: (
            str(r.get("discord_user", "")).casefold(),
            str(r.get("account_type", "")).casefold(),
        )
    )
    unreg_rows.sort(
        key=lambda r: (str(r.get("governor_name", "")).casefold(), str(r.get("governor_id", "")))
    )
    mw_rows.sort(key=lambda r: str(r.get("discord_user", "")).casefold())

    reg_headers = [
        "discord_id",
        "discord_id_excel",
        "discord_user",
        "account_type",
        "governor_id",
        "governor_id_excel",
        "governor_name",
        "roles",
        "top_role",
    ]
    unreg_headers = [
        "governor_id",
        "governor_id_excel",
        "governor_name",
        "alliance",
        "power_rank",
        "power",
        "kill_points",
        "deads",
        "city_hall",
        "location",
        "scan_date",
    ]
    mw_headers = ["discord_id", "discord_id_excel", "discord_user", "roles", "top_role"]

    return {
        "registered_accounts.csv": to_csv_bytes(reg_headers, registered_rows),
        "unregistered_current_governors.csv": to_csv_bytes(unreg_headers, unreg_rows),
        "members_without_registration.csv": to_csv_bytes(mw_headers, mw_rows),
    }


def export_registration_audit_xlsx_bytes(
    registry: dict,
    members_info: dict[str, dict[str, str]],
    sql_rows: list[dict[str, object]],
) -> io.BytesIO:
    """
    Produce a single XLSX workbook with three sheets: Registered, Unregistered, MembersWithoutRegistration.
    Returns a BytesIO.
    """
    _ensure_openpyxl()
    from openpyxl import Workbook
    from openpyxl.styles import Alignment

    files = export_registration_audit_files(registry, members_info, sql_rows)
    # files contains CSV BytesIOs; convert each CSV back to rows and write into separate sheets
    wb = Workbook()
    # Registered sheet
    reg_sheet = wb.active
    reg_sheet.title = "registered_accounts"
    reg_csv = files["registered_accounts.csv"].getvalue().decode("utf-8-sig")
    reg_rows = list(csv.reader(io.StringIO(reg_csv)))
    for r in reg_rows:
        reg_sheet.append(r)
    # Unregistered
    unreg_sheet = wb.create_sheet(title="unregistered_current_governors")
    unreg_csv = files["unregistered_current_governors.csv"].getvalue().decode("utf-8-sig")
    unreg_rows = list(csv.reader(io.StringIO(unreg_csv)))
    for r in unreg_rows:
        unreg_sheet.append(r)
    # Members without registration
    mw_sheet = wb.create_sheet(title="members_without_registration")
    mw_csv = files["members_without_registration.csv"].getvalue().decode("utf-8-sig")
    mw_rows = list(csv.reader(io.StringIO(mw_csv)))
    for r in mw_rows:
        mw_sheet.append(r)

    # Force text format on id columns for each sheet where applicable
    text_cols_by_sheet = {
        "registered_accounts": [
            1,
            2,
            5,
            6,
        ],  # discord_id, discord_id_excel, governor_id, governor_id_excel
        "unregistered_current_governors": [1, 2],  # governor_id, governor_id_excel
        "members_without_registration": [1, 2],  # discord_id, discord_id_excel
    }
    for sheet in wb.worksheets:
        cols = text_cols_by_sheet.get(sheet.title, [])
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for col_idx in cols:
                cell = row[col_idx - 1]
                if cell.value is None:
                    continue
                cell.value = str(cell.value)
                try:
                    cell.number_format = "@"
                    cell.alignment = Alignment(horizontal="left")
                except Exception:
                    pass

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
