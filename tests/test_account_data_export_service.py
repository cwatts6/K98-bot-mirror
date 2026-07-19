from __future__ import annotations

import csv
from dataclasses import replace
from datetime import UTC, datetime
from pathlib import Path
import threading

from openpyxl import load_workbook
import pandas as pd
import pytest

from player_self_service.account_data_export_contract import AccountDataOutputKind
from player_self_service.accounts_models import (
    AccountMetricTotal,
    AccountPortfolioRow,
    AccountsPortfolioPayload,
)
from services import account_data_export_service as service

GENERATED = datetime(2026, 7, 16, 9, 30, tzinfo=UTC)


def _portfolio(user_id: int = 42) -> AccountsPortfolioPayload:
    inventory_as_of = datetime(2026, 7, 15, 8, 0, tzinfo=UTC)
    rows = (
        AccountPortfolioRow(
            slot="Main",
            role="Main",
            registered_name="=Main",
            current_governor_name="Main",
            governor_id=111,
            power=100,
            data_state="CURRENT",
            last_governor_scan=GENERATED,
            inventory_as_of=inventory_as_of,
        ),
        AccountPortfolioRow(
            slot="Alt 1",
            role="Alt",
            registered_name="Alt",
            current_governor_name="Alt",
            governor_id=222,
            power=50,
            data_state="CURRENT",
            last_governor_scan=GENERATED,
            inventory_as_of=inventory_as_of,
        ),
    )
    metric = AccountMetricTotal(value=150, reporting_count=2, expected_count=2)
    return AccountsPortfolioPayload(
        discord_user_id=user_id,
        state="READY",
        rows=rows,
        linked_count=2,
        main_row=rows[0],
        role_counts=(("Main", 1), ("Alt", 1), ("Farm", 0)),
        power=metric,
        troop_power=metric,
        t4_t5_kills=metric,
        rss_total=metric,
        insight="Ready",
        refreshed_at_utc=GENERATED,
        latest_scan_date=GENERATED,
    )


def _history(days: int = 100) -> pd.DataFrame:
    rows = []
    dates = pd.date_range(end="2026-07-16", periods=days, freq="D")
    for governor_id, name in ((111, "=Main"), (222, "+Alt")):
        for index, source_date in enumerate(dates):
            rows.append(
                {
                    "GovernorID": governor_id,
                    "GovernorName": name,
                    "Alliance": "@K98",
                    "AsOfDate": source_date,
                    "Power": 100 + index,
                    "PowerDelta": -5,
                }
            )
    return pd.DataFrame(rows)


def _install_authority(monkeypatch, *, user_id: int = 42) -> list[tuple[int, ...]]:
    resolution = service.governor_account_service.summarize_accounts(
        {
            "Main": {"GovernorID": "111", "GovernorName": "Main"},
            "Alt 1": {"GovernorID": "222", "GovernorName": "Alt"},
        }
    )

    async def fake_resolution(requested_user_id: int):
        assert requested_user_id == user_id
        return resolution

    async def fake_portfolio(requested_user_id: int, supplied_resolution, **kwargs):
        assert requested_user_id == user_id
        assert supplied_resolution is resolution
        assert kwargs["refreshed_at_utc"] == GENERATED
        return _portfolio(user_id)

    monkeypatch.setattr(
        service.governor_account_service, "get_account_summary_for_user", fake_resolution
    )
    monkeypatch.setattr(
        service.accounts_service, "build_accounts_portfolio_from_resolution", fake_portfolio
    )
    calls: list[tuple[int, ...]] = []
    return calls


def _install_temp_dir(monkeypatch, tmp_path: Path) -> Path:
    target = tmp_path / "account-data"

    def fake_mkdtemp(*, prefix: str):
        assert prefix == "k98_account_data_"
        target.mkdir()
        return str(target)

    monkeypatch.setattr(service.tempfile, "mkdtemp", fake_mkdtemp)
    return target


@pytest.mark.asyncio
async def test_snapshot_revalidates_once_and_does_not_query_stats(monkeypatch, tmp_path) -> None:
    _install_authority(monkeypatch)
    _install_temp_dir(monkeypatch, tmp_path)
    monkeypatch.setattr(
        service.stats_export_dal,
        "fetch_daily_player_export",
        lambda _ids: pytest.fail("snapshot must not query Stats history"),
    )

    outcome = await service.build_account_data_export(
        discord_user_id=42,
        display_name="Tester",
        requested_kind=AccountDataOutputKind.CURRENT_SNAPSHOT,
        requested_days=360,
        generated_at_utc=GENERATED,
    )

    assert outcome.status == "ok"
    assert outcome.export_file is not None
    assert outcome.export_file.filename == "me_account_summary_42_20260716_093000.csv"
    assert outcome.export_file.metadata.snapshot_row_count == 2
    assert outcome.export_file.metadata.history_row_count is None
    assert outcome.export_file.metadata.stats_freshness is None
    assert outcome.export_file.metadata.governor_scan_freshness == GENERATED
    assert outcome.export_file.metadata.inventory_reporting_count == 2
    assert outcome.export_file.metadata.inventory_expected_count == 2
    rows = list(
        csv.reader(outcome.export_file.file_path.read_text(encoding="utf-8-sig").splitlines())
    )
    assert len(rows[0]) == 29
    assert rows[1][2] == "'=Main"
    service.cleanup_export_file(outcome.export_file)
    assert not outcome.export_file.temp_dir.exists()


@pytest.mark.asyncio
async def test_raw_history_filters_before_metadata_and_writes_formula_safe_text(
    monkeypatch, tmp_path
) -> None:
    calls = _install_authority(monkeypatch)
    _install_temp_dir(monkeypatch, tmp_path)

    def fake_fetch(governor_ids):
        calls.append(tuple(governor_ids))
        return _history()

    monkeypatch.setattr(service.stats_export_dal, "fetch_daily_player_export", fake_fetch)

    outcome = await service.build_account_data_export(
        discord_user_id=42,
        display_name="Test User",
        requested_kind="raw_history",
        requested_days=90,
        generated_at_utc=GENERATED,
    )

    assert calls == [(111, 222)]
    assert outcome.status == "ok"
    assert outcome.export_file is not None
    assert outcome.export_file.filename == "stats_Test_User_20260716_093000.csv"
    metadata = outcome.export_file.metadata
    assert metadata.history_row_count == 180
    assert metadata.window_start.isoformat() == "2026-04-18"
    assert metadata.window_end.isoformat() == "2026-07-16"
    with outcome.export_file.file_path.open(encoding="utf-8", newline="") as stream:
        rows = list(csv.DictReader(stream))
    assert len(rows) == 180
    assert rows[0]["GovernorName"] == "'=Main"
    assert rows[0]["Alliance"] == "'@K98"
    assert rows[0]["PowerDelta"] == "-5"


@pytest.mark.asyncio
async def test_full_workbook_uses_filtered_rows_and_locked_sheet_order(
    monkeypatch, tmp_path
) -> None:
    _install_authority(monkeypatch)
    _install_temp_dir(monkeypatch, tmp_path)
    monkeypatch.setattr(
        service.stats_export_dal, "fetch_daily_player_export", lambda _ids: _history()
    )

    outcome = await service.build_account_data_export(
        discord_user_id=42,
        display_name="Test User",
        requested_kind="full_workbook",
        requested_days=30,
        generated_at_utc=GENERATED,
    )

    assert outcome.status == "ok"
    assert outcome.export_file is not None
    assert outcome.export_file.metadata.history_row_count == 60
    assert outcome.export_file.metadata.stats_freshness.isoformat() == "2026-07-16"
    assert outcome.export_file.metadata.governor_scan_freshness == GENERATED
    assert outcome.export_file.metadata.inventory_reporting_count == 2
    workbook = load_workbook(outcome.export_file.file_path, read_only=True, data_only=False)
    assert workbook.sheetnames[:3] == ["ACCOUNT_SUMMARY", "README", "ALL_DAILY"]
    assert workbook["ACCOUNT_SUMMARY"].max_column == 29
    assert workbook["ALL_DAILY"].max_row == 61
    workbook.close()
    service.cleanup_export_file(outcome.export_file)


@pytest.mark.asyncio
async def test_download_uses_links_re_resolved_at_execution(monkeypatch, tmp_path) -> None:
    resolution = service.governor_account_service.summarize_accounts(
        {"Main": {"GovernorID": "222", "GovernorName": "Relinked"}}
    )

    async def fake_resolution(_user_id: int):
        return resolution

    async def fake_portfolio(_user_id: int, supplied_resolution, **_kwargs):
        assert supplied_resolution is resolution
        original = _portfolio()
        relinked_row = replace(original.rows[1], slot="Main", role="Main")
        return replace(original, rows=(relinked_row,), linked_count=1, main_row=relinked_row)

    seen: list[tuple[int, ...]] = []

    def fake_fetch(governor_ids):
        seen.append(tuple(governor_ids))
        return _history(2).loc[lambda frame: frame["GovernorID"] == 222].copy()

    monkeypatch.setattr(
        service.governor_account_service, "get_account_summary_for_user", fake_resolution
    )
    monkeypatch.setattr(
        service.accounts_service, "build_accounts_portfolio_from_resolution", fake_portfolio
    )
    monkeypatch.setattr(service.stats_export_dal, "fetch_daily_player_export", fake_fetch)
    _install_temp_dir(monkeypatch, tmp_path)

    outcome = await service.build_account_data_export(
        discord_user_id=42,
        display_name="Test User",
        requested_kind="raw_history",
        requested_days=30,
        generated_at_utc=GENERATED,
    )

    assert outcome.status == "ok"
    assert seen == [(222,)]
    assert outcome.export_file is not None
    assert outcome.export_file.metadata.authorised_governor_count == 1
    service.cleanup_export_file(outcome.export_file)


@pytest.mark.asyncio
async def test_service_rejects_unauthorised_history_rows(monkeypatch) -> None:
    _install_authority(monkeypatch)
    frame = _history(1)
    frame.loc[len(frame)] = {
        "GovernorID": 999,
        "GovernorName": "Other",
        "Alliance": "K98",
        "AsOfDate": pd.Timestamp("2026-07-16"),
        "Power": 1,
        "PowerDelta": 0,
    }
    monkeypatch.setattr(service.stats_export_dal, "fetch_daily_player_export", lambda _ids: frame)

    outcome = await service.build_account_data_export(
        discord_user_id=42,
        display_name="Tester",
        requested_kind="raw_history",
        requested_days=30,
        generated_at_utc=GENERATED,
    )

    assert outcome.status == "data_error"
    assert outcome.export_file is None


@pytest.mark.asyncio
async def test_builder_failure_cleans_owned_temp_directory(monkeypatch, tmp_path) -> None:
    _install_authority(monkeypatch)
    target = _install_temp_dir(monkeypatch, tmp_path)
    monkeypatch.setattr(
        service.stats_export_dal, "fetch_daily_player_export", lambda _ids: _history(2)
    )

    def fail_builder(*_args, out_path, **_kwargs):
        Path(out_path).write_text("partial", encoding="utf-8")
        raise RuntimeError("builder failed")

    monkeypatch.setattr(service, "build_account_data_history_csv", fail_builder)

    outcome = await service.build_account_data_export(
        discord_user_id=42,
        display_name="Tester",
        requested_kind="raw_history",
        requested_days=30,
        generated_at_utc=GENERATED,
    )

    assert outcome.status == "generation_error"
    assert not target.exists()


@pytest.mark.asyncio
async def test_builder_cancellation_waits_for_writer_then_cleans_temp_directory(
    monkeypatch, tmp_path
) -> None:
    _install_authority(monkeypatch)
    target = _install_temp_dir(monkeypatch, tmp_path)
    monkeypatch.setattr(
        service.stats_export_dal, "fetch_daily_player_export", lambda _ids: _history(2)
    )
    opened = threading.Event()
    release = threading.Event()
    finished = threading.Event()

    def blocking_builder(*_args, out_path, **_kwargs):
        try:
            with Path(out_path).open("wb") as handle:
                handle.write(b"partial private export")
                handle.flush()
                opened.set()
                assert release.wait(timeout=10)
                handle.write(b" completed")
        finally:
            finished.set()

    monkeypatch.setattr(service, "build_account_data_history_csv", blocking_builder)
    task = service.asyncio.create_task(
        service.build_account_data_export(
            discord_user_id=42,
            display_name="Tester",
            requested_kind="raw_history",
            requested_days=30,
            generated_at_utc=GENERATED,
        )
    )
    assert await service.asyncio.to_thread(opened.wait, 10)
    task.cancel()

    with pytest.raises(service.asyncio.CancelledError):
        await task

    assert target.exists()
    release.set()
    assert await service.asyncio.to_thread(finished.wait, 10)
    for _ in range(100):
        if not target.exists():
            break
        await service.asyncio.sleep(0.01)

    assert not target.exists()


@pytest.mark.asyncio
async def test_invalid_output_and_days_fail_before_registry(monkeypatch) -> None:
    monkeypatch.setattr(
        service.governor_account_service,
        "get_account_summary_for_user",
        lambda _user_id: pytest.fail("invalid request must not load the registry"),
    )

    invalid_kind = await service.build_account_data_export(
        discord_user_id=42,
        display_name="Tester",
        requested_kind="GoogleSheets",
        requested_days=90,
    )
    invalid_days = await service.build_account_data_export(
        discord_user_id=42,
        display_name="Tester",
        requested_kind="full_workbook",
        requested_days=91,
    )

    assert invalid_kind.status == "invalid_request"
    assert invalid_days.status == "invalid_request"
