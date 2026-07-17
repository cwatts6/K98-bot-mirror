# tests/test_stats_export.py
"""Unit tests for stats export functionality."""

from datetime import UTC, date as date_type, datetime
import os
import tempfile
import warnings
import zipfile

from openpyxl import load_workbook
import pandas as pd
import pytest

from player_self_service.account_data_export_contract import (
    AccountDataExportMetadata,
    AccountDataOutputKind,
)
from player_self_service.accounts_export import CSV_COLUMNS
from player_self_service.accounts_models import (
    AccountMetricTotal,
    AccountPortfolioRow,
    AccountsPortfolioPayload,
)
from stats_exporter import (
    _calc_period,
    _clean_text,
    _safe_sheet_name,
    build_account_data_workbook,
    build_user_stats_excel,
)


@pytest.fixture
def sample_daily_data():
    """Generate sample daily stats data matching vDaily_PlayerExport schema."""
    dates = pd.date_range(end=datetime.now(), periods=90, freq="D")
    data = []

    for i, date in enumerate(dates):
        data.append(
            {
                "GovernorID": 123456,
                "GovernorName": "TestPlayer",
                "Alliance": "K98",
                "AsOfDate": date,
                # Core metrics
                "Power": 50000000 + (i * 100000),
                "PowerDelta": 100000,
                "TroopPower": 40000000 + (i * 80000),
                "TroopPowerDelta": 80000,
                "KillPoints": 1000000 + (i * 10000),
                "KillPointsDelta": 10000,
                "Deads": 500000 + (i * 5000),
                "DeadsDelta": 5000,
                # Activity
                "RSS_Gathered": 10000000000 + (i * 100000000),
                "RSS_GatheredDelta": 100000000,
                "RSSAssist": 5000000000 + (i * 50000000),
                "RSSAssistDelta": 50000000,
                "Helps": 10000 + (i * 100),
                "HelpsDelta": 100,
                "BuildingMinutes": 500,
                "TechDonations": 200,
                # Forts
                "FortsTotal": 10 + (i // 7),  # Increment weekly
                "FortsLaunched": 5 + (i // 7),
                "FortsJoined": 5 + (i // 7),
                # AOO
                "AOOJoined": min(i // 7, 10),  # Max 10 events
                "AOOJoinedDelta": 1 if i % 7 == 0 else 0,
                "AOOWon": min(i // 14, 5),  # Max 5 wins
                "AOOWonDelta": 1 if i % 14 == 0 else 0,
                "AOOAvgKill": 1000000 + (i * 1000),
                "AOOAvgKillDelta": 1000,
                "AOOAvgDead": 500000 + (i * 500),
                "AOOAvgDeadDelta": 500,
                "AOOAvgHeal": 800000 + (i * 800),
                "AOOAvgHealDelta": 800,
                # Detailed metrics
                "T4_Kills": 50000 + (i * 500),
                "T4_KillsDelta": 500,
                "T5_Kills": 30000 + (i * 300),
                "T5_KillsDelta": 300,
                "T4T5_Kills": 80000 + (i * 800),
                "T4T5_KillsDelta": 800,
                "HealedTroops": 100000 + (i * 1000),
                "HealedTroopsDelta": 1000,
                "RangedPoints": 200000 + (i * 2000),
                "RangedPointsDelta": 2000,
                "HighestAcclaim": 50,
                "HighestAcclaimDelta": 0,
                "AutarchTimes": 2,
                "AutarchTimesDelta": 0,
            }
        )

    return pd.DataFrame(data)


def test_calc_period_basic(sample_daily_data):
    """Test period calculation for standard metrics."""
    df = sample_daily_data
    end_date = pd.Timestamp(df["AsOfDate"].max())
    start_date = end_date - pd.Timedelta(days=30)

    result = _calc_period(df, start_date, end_date)

    # Verify snapshots are from end of period
    assert result["PowerEnd"] > 0
    assert result["TroopPowerEnd"] > 0

    # Verify deltas are sums
    assert result["PowerDelta"] == 100000 * 31  # 31 days
    assert result["KillPointsDelta"] == 10000 * 31


def test_calc_period_aoo_metrics(sample_daily_data):
    """Test AOO metric calculation."""
    df = sample_daily_data
    end_date = pd.Timestamp(df["AsOfDate"].max())
    start_date = end_date - pd.Timedelta(days=30)

    result = _calc_period(df, start_date, end_date)

    # AOO metrics should be snapshots (cumulative)
    assert "AOOJoinedEnd" in result
    assert "AOOWonEnd" in result
    assert result["AOOJoinedEnd"] >= 0
    assert result["AOOWonEnd"] >= 0


def test_calc_period_detailed_kills(sample_daily_data):
    """Test detailed kill metric calculation."""
    df = sample_daily_data
    end_date = pd.Timestamp(df["AsOfDate"].max())
    start_date = end_date - pd.Timedelta(days=7)

    result = _calc_period(df, start_date, end_date)

    # Verify detailed metrics present
    assert "T4_KillsEnd" in result
    assert "T5_KillsEnd" in result
    assert "T4T5_KillsEnd" in result
    assert "HealedTroopsEnd" in result

    # Verify deltas
    assert result["T4_KillsDelta"] == 500 * 8  # 8 days
    assert result["T5_KillsDelta"] == 300 * 8


def test_clean_text():
    """Test text cleaning function."""
    assert _clean_text("  Test  Player  ") == "Test Player"
    assert _clean_text(None) == ""
    assert _clean_text("Test\nPlayer") == "Test Player"


def test_safe_sheet_name():
    """Test Excel sheet name sanitization."""
    # Test forbidden characters
    assert "[" not in _safe_sheet_name("Test[1]", "fallback")
    assert ":" not in _safe_sheet_name("Test:Player", "fallback")

    # Test length limit
    long_name = "A" * 50
    result = _safe_sheet_name(long_name, "fallback", max_len=31)
    assert len(result) <= 31

    # Test fallback
    result = _safe_sheet_name("", "MyFallback")
    assert result == "MyFallback"


def test_build_excel_basic(sample_daily_data):
    """Test basic Excel generation."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        # Should not raise
        build_user_stats_excel(sample_daily_data, None, out_path=tmp_path)

        # Verify file exists and has content
        assert os.path.exists(tmp_path)
        assert os.path.getsize(tmp_path) > 0

        # Verify can be read back
        df_test = pd.read_excel(tmp_path, sheet_name="ALL_DAILY")
        assert not df_test.empty
        assert "GovernorID" in df_test.columns
        assert "AOOJoined" in df_test.columns
        assert "T4_Kills" in df_test.columns

    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


def test_build_excel_multiple_governors(sample_daily_data):
    """Test Excel with multiple governors."""
    # Add second governor
    df2 = sample_daily_data.copy()
    df2["GovernorID"] = 789012
    df2["GovernorName"] = "Player2"

    combined = pd.concat([sample_daily_data, df2], ignore_index=True)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        build_user_stats_excel(combined, None, out_path=tmp_path)

        # Verify both governors present
        df_test = pd.read_excel(tmp_path, sheet_name="ALL_DAILY")
        assert 123456 in df_test["GovernorID"].values
        assert 789012 in df_test["GovernorID"].values

    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


def test_deprecated_columns_removed(sample_daily_data):
    """Verify deprecated columns are not in output."""
    # Add deprecated columns to source
    df = sample_daily_data.copy()
    df["TechPower"] = 1000000
    df["TechPowerDelta"] = 1000

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        build_user_stats_excel(df, None, out_path=tmp_path)

        df_test = pd.read_excel(tmp_path, sheet_name="ALL_DAILY")

        # Verify deprecated columns removed
        assert "TechPower" not in df_test.columns
        assert "TechPowerDelta" not in df_test.columns

    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


def test_no_dkp_in_export():
    """Verify DKP/targets removed from export per requirements."""
    # This test verifies the signature and that targets param is ignored
    df = pd.DataFrame(
        {
            "GovernorID": [123],
            "GovernorName": ["Test"],
            "Alliance": ["K98"],
            "AsOfDate": [datetime.now()],
            "Power": [50000000],
            "PowerDelta": [0],
        }
    )

    # Fake targets dataframe (should be ignored)
    df_targets = pd.DataFrame(
        {
            "GovernorID": [123],
            "DKP_TARGET": [100],
            "DKP_SCORE": [80],
        }
    )

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        # Targets should be ignored (set to None internally)
        build_user_stats_excel(df, df_targets, out_path=tmp_path)

        # Verify no DKP columns in output
        df_test = pd.read_excel(tmp_path, sheet_name="ALL_DAILY")
        assert "DKP_TARGET" not in df_test.columns
        assert "DKP_SCORE" not in df_test.columns

    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


def test_account_data_workbook_locked_contract(sample_daily_data, tmp_path):
    first = sample_daily_data.tail(30).copy()
    first["GovernorID"] = 111
    first["GovernorName"] = "=Unsafe governor"
    first["Alliance"] = "+Unsafe alliance"
    first["FortsTotal"] = 1
    first["FortsLaunched"] = 2
    first["FortsJoined"] = 3
    second = first.iloc[:1].copy()
    second["GovernorID"] = 222
    second["GovernorName"] = "=Unsafe governor"
    history = pd.concat([first, second], ignore_index=True)

    rows = (
        AccountPortfolioRow(
            slot="Main",
            role="Main",
            registered_name="=Formula",
            current_governor_name="Same very long governor name that must be truncated safely",
            governor_id=111,
            data_state="CURRENT",
        ),
        AccountPortfolioRow(
            slot="Alt 1",
            role="Alt",
            registered_name="Alt",
            current_governor_name="Same very long governor name that must be truncated safely",
            governor_id=222,
            data_state="CURRENT",
        ),
        AccountPortfolioRow(
            slot="Farm 1",
            role="Farm",
            registered_name="Sparse",
            current_governor_name="O'Brien / sparse",
            governor_id=333,
            data_state="NO DATA",
        ),
    )
    metric = AccountMetricTotal(value=0, reporting_count=0, expected_count=3)
    generated = datetime(2026, 7, 16, 9, 30, tzinfo=UTC)
    portfolio = AccountsPortfolioPayload(
        discord_user_id=42,
        state="REVIEW",
        rows=rows,
        linked_count=3,
        main_row=rows[0],
        role_counts=(("Main", 1), ("Alt", 1), ("Farm", 1)),
        power=metric,
        troop_power=metric,
        t4_t5_kills=metric,
        rss_total=metric,
        insight="Review",
        refreshed_at_utc=generated,
    )
    metadata = AccountDataExportMetadata(
        output_kind=AccountDataOutputKind.FULL_WORKBOOK,
        generated_at_utc=generated,
        authorised_governor_count=3,
        snapshot_row_count=3,
        history_row_count=len(history.index),
        requested_days=30,
        window_start=date_type(2026, 6, 17),
        window_end=date_type(2026, 7, 16),
        written_start=pd.Timestamp(first["AsOfDate"].min()).date(),
        written_end=pd.Timestamp(first["AsOfDate"].max()).date(),
        stats_freshness=pd.Timestamp(first["AsOfDate"].max()).date(),
        governor_scan_freshness=None,
        inventory_oldest=None,
        inventory_latest=None,
        inventory_reporting_count=0,
        inventory_expected_count=3,
    )
    target = tmp_path / "account-data.xlsx"

    build_account_data_workbook(
        history,
        portfolio=portfolio,
        governor_ids=(111, 222, 333),
        metadata=metadata,
        out_path=target,
    )

    with zipfile.ZipFile(target) as archive:
        assert archive.testzip() is None
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        workbook = load_workbook(target, data_only=False)
    assert all("Sparkline Group extension is not supported" in str(item.message) for item in caught)
    assert workbook.sheetnames[:3] == ["ACCOUNT_SUMMARY", "README", "ALL_DAILY"]
    account_sheets = workbook.sheetnames[3:]
    assert len(account_sheets) == 3
    assert len({name.casefold() for name in account_sheets}) == 3
    assert all(len(name) <= 31 for name in account_sheets)

    summary = workbook["ACCOUNT_SUMMARY"]
    assert tuple(cell.value for cell in summary[1]) == CSV_COLUMNS
    assert summary.max_column == 29
    assert summary["C2"].value == "'=Formula"
    assert summary["D2"].hyperlink is not None
    assert account_sheets[0] in summary["D2"].hyperlink.location
    assert summary["D4"].hyperlink is not None
    assert summary["D4"].hyperlink.location == "'O''Brien  sparse-333'!A1"

    all_daily = workbook["ALL_DAILY"]
    assert all_daily.max_row == len(history.index) + 1
    assert all_daily["B2"].data_type == "s"
    assert str(all_daily["B2"].value).startswith("'")
    assert str(all_daily["C2"].value).startswith("'")

    first_sheet = workbook[account_sheets[0]]
    assert first_sheet["B16"].value == 30
    assert first_sheet["C16"].value == 60
    assert first_sheet["D16"].value == 90
    assert first_sheet["B19"].value == "RSS_Gathered"
    assert first_sheet["B20"].value == "GovernorName"
    assert len(first_sheet._charts) == 5
    sparse_sheet = workbook[account_sheets[2]]
    assert "No Stats history" in sparse_sheet["A4"].value
    assert not sparse_sheet._charts


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
