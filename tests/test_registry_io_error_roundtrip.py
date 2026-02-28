# tests/test_registry_io_error_roundtrip.py
import csv
import io

import pytest

import registry_io


def make_error_rows():
    return [
        {
            "row": 2,
            "discord_id": "123",
            "account_type": "Main",
            "governor_id": "100001",
            "error": "Invalid governor",
        },
        {
            "row": 3,
            "discord_id": "456",
            "account_type": "Alt 1",
            "governor_id": "100002",
            "error": "Duplicate governor",
        },
    ]


def test_build_error_csv_bytes_contains_rows():
    rows = make_error_rows()
    buf = registry_io.build_error_csv_bytes(rows)
    assert isinstance(buf, io.BytesIO)
    text = buf.getvalue().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    out = list(reader)
    assert len(out) == 2
    # headers present and first discord_id matches
    assert "discord_id" in reader.fieldnames
    assert out[0]["discord_id"] == "123"
    assert out[1]["error"] == "Duplicate governor"


def test_build_error_xlsx_bytes_roundtrip():
    # Skip if openpyxl not available
    pytest.importorskip("openpyxl")

    rows = make_error_rows()
    xbuf = registry_io.build_error_xlsx_bytes(rows)
    assert isinstance(xbuf, io.BytesIO)

    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(xbuf.getvalue()), read_only=True)
    ws = wb.active
    # header row (values_only returns raw values)
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip().lower() for h in header_row]
    assert "discord_id" in headers
    # data rows
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 2
    idx = headers.index("discord_id")
    assert data_rows[0][idx] == "123"
    assert data_rows[1][headers.index("error")] == "Duplicate governor"
